from io import BytesIO

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def format_category(category_type, category_value):
    if category_value is None:
        return "-"

    if category_type == "bedrooms":
        if category_value == 0:
            return "Studio"
        if category_value == 1:
            return "1 Bedroom"
        return f"{category_value} Bedrooms"

    return str(category_value)


def auto_fit_columns(sheet, min_width=12, max_width=40):
    for column_cells in sheet.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)

        adjusted_width = max(min_width, min(max_length + 2, max_width))
        sheet.column_dimensions[column_letter].width = adjusted_width


def generate_summary_excel(summary_rows):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Summary"

    headers = [
        "Source File",
        "Developer",
        "Project",
        "Category",
        "Starting Price",
        "Starting Area (m2)",
    ]

    sheet.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in summary_rows:
        sheet.append([
            row.get("source_file") or "-",
            row.get("developer_name") or "-",
            row.get("project_name") or "-",
            format_category(row.get("category_type"), row.get("category_value")),
            row.get("starting_price"),
            row.get("starting_area_m2"),
        ])

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in range(2, sheet.max_row + 1):
        sheet[f"E{row}"].number_format = "#,##0.00"
        sheet[f"F{row}"].number_format = "#,##0.00"

    auto_fit_columns(sheet)

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer